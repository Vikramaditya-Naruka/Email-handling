#!/usr/bin/env python
# coding: utf-8

# In[ ]:





# Email handling in Python
# 
# 1.) Sending Emails:
# 
# Python provides the smtplib library for sending emails.
# With smtplib, you can connect to an SMTP (Simple Mail Transfer Protocol) server 
# and send emails using your email provider's settings.
# 
# 2.) Receiving Emails:
# 
# For receiving emails, you can use the imaplib library.
# IMAP (Internet Message Access Protocol) allows you to retrieve and manage emails stored on a server.
# 
# 3.) Parsing Email Content:
# 
# To parse email content, you can use the email library,
# which provides classes and functions for working with email messages,
# including headers, attachments, and body content.
# 
# 4.) Working with Attachments:
# 
# You can also handle email attachments using the 'email' library. 
# Attachments are usually parts of a multipart message.
# 
# 5.)Decoding Headers:
# 
# The 'email.header' module helps decode headers that might contain non-ASCII characters.

# In[ ]:





# In[ ]:





# IMPORTANT Notes:
# 
# Security ===> Storing passwords directly in code is not secure. Use "App Passwords" or other secure methods.
# 
# Less Secure Apps ===> For Gmail, you might need to allow "Less Secure Apps" in your account settings.
# 
# App Passwords ===> If you have two-factor authentication enabled, use "App Passwords" for added security.
# 
# Environment Variables ===> For sensitive data like passwords, use environment variables to keep them safe.

# 

# In[ ]:





# In[ ]:


'''
i have an account that has arround 5000 emails and i want to run RPA on it and 
extract the following information for each mail .

1.)sender Email id
2.) Mail Subject
3.) Attachment Yes/No
'''




import imaplib
from email.header import decode_header
import email
import openpyxl
import getpass
import os


# Replace with your email account information
EMAIL = "vaishalisharmay93@gmail.com"
PASSWORD = getpass.getpass("Enter your Email Password : ")

# Connect to the IMAP server
mail = imaplib.IMAP4_SSL("imap.gmail.com")

# Log in to your email account
mail.login(EMAIL, PASSWORD)

# Select the mailbox you want to extract emails from (e.g., "INBOX")
mail.select("INBOX")
# Search for all emails
status, email_ids = mail.search(None, "ALL")

# Create an Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["Sender Email", "Mail Subject", "Attachment"])


                        
# Iterate through email IDs and fetch details

for email_id in email_ids[0].split():
    status, msg_data = mail.fetch(email_id, "(BODY.PEEK[])")
    for response_part in msg_data:
        if isinstance(response_part, tuple):
            msg = response_part[1]
            email_message = email.message_from_bytes(msg)
            sender = decode_header(email_message['From'])[0][0]
            subject = decode_header(email_message['Subject'])[0][0]

            has_attachments = False  # Initialize flag for attachments
        
            # Check if email has attachments
            if email_message.is_multipart():
                for part in email_message.walk():
                    content_type = part.get_content_type()
                    if "attachment" in content_type:
                        has_attachments = True
                        
                        # Get attachment filename
                        filename = part.get_filename()
                        
                        # Save the attachment to a local file
                        if filename:
                            attachment_data = part.get_payload(decode=True)
                            with open(filename, "wb") as f:
                                f.write(attachment_data)
            
            # Print email details and attachment status
            print("Sender:", sender)
            print("Subject:", subject)
            print("Attachments:", "Yes" if has_attachments else "No")
            print("-" * 50)
            
            #Append extracted information to the Excel sheet
            
            sheet.append([sender, subject,  has_attachments])

# Save the workbook to a file
workbook.save("emaist.xlsx")
 

# Logout and close the connection
mail.logout()


# # also download the attachment

# In[ ]:


import imaplib
import openpyxl
import email
from email.header import decode_header
import getpass
import os

# Email credentials
email_user = "vaishalisharmay93@gmail.com"
email_password = getpass.getpass("Enter your email password: ")

# Connect to the mail server
mail = imaplib.IMAP4_SSL("imap.gmail.com")
mail.login(email_user, email_password)
mail.select("inbox")

# Search for all emails
status, email_ids = mail.search(None, "ALL")
email_ids = email_ids[0].split()

# Create an Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.append(["Sender Email", "Mail Subject", "Attachment"])

# Create a directory to save attachments
attachment_dir = "attachments"
os.makedirs(attachment_dir, exist_ok=True)

# Loop through emails
for email_id in email_ids:
    status, msg_data = mail.fetch(email_id, "(RFC822)")
    msg = email.message_from_bytes(msg_data[0][1])

    # Extract sender email
    sender = msg["From"]

    # Extract subject
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        subject = subject.decode(encoding or "utf-8")

    # Check for attachments
    has_attachments = any(part.get_content_disposition() for part in msg.walk())
    attachment = "Yes" if has_attachments else "No"

    # Download attachments if available
    attachment_filenames = []
    if has_attachments:
        for part in msg.walk():
            if part.get_content_disposition():
                filename = part.get_filename()
                if filename:
                    attachment_filenames.append(filename)
                    filepath = os.path.join(attachment_dir, filename)
                    with open(filepath, "wb") as f:
                        f.write(part.get_payload(decode=True))

    # Append extracted information to the Excel sheet
    sheet.append([sender, subject, attachment])

    # Add attachment filenames to the Excel sheet
    attachment_info = ", ".join(attachment_filenames) if attachment_filenames else "No attachments"
    sheet.cell(row=sheet.max_row, column=3).value = attachment_info

    print("=" * 40)

# Save the workbook to a file
workbook.save("e_information.xlsx")

# Disconnect from the mail server
mail.logout()


# # for sending message

# In[ ]:


# message sending on mails

import smtplib
from email.message import EmailMessage

# Create an EmailMessage object
msg = EmailMessage()
msg.set_content(" this is the email content!")

msg['Subject'] = 'dil mange more'
msg['From'] = 'vaishalisharmay93@gmail.com'
msg['To'] = 'vikramadityasinghn@gmail.com'

# Connect to the SMTP server and send the email
with smtplib.SMTP('smtp.gmail.com', 587) as server:
    server.starttls()  # Secure the connection
    server.login('vaishalisharmay93@gmail.com', 'dwaswubatxepcpkb')
    server.send_message(msg)
    
#server.quit()    


# In[ ]:


# message sending on mails

import smtplib

def send_email(subject, body, to_email, from_email, password):
    msg = f"Subject: {subject}\n\n{body}"

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()  # Secure the connection
            server.login(from_email, password)
            server.sendmail(from_email, to_email, msg)
        print("Email sent successfully!")
    except Exception as e:
        print("An error occurred:", e)

if __name__ == "__main__":
    your_email = input("Enter your Gmail email address: ")
    your_password = input("Enter your Gmail password: ")
    recipient_email = input("Enter recipient's email address: ")
    email_subject = input("Enter email subject: ")
    email_body = input("Enter email body: ")

    send_email(email_subject, email_body, recipient_email, your_email, your_password)
    

#server.quit()


# In[ ]:


import smtplib

'''
1.)smtplib module is a built-in Python library that provides a way to send emails using the Simple Mail Transfer Protocol(SMTP).
SMTP is a protocol for sending emails between servers. 
The smtplib module simplifies the process of sending emails from your Python script.
'''
from email.message import EmailMessage

msg = EmailMessage()
msg['Subject'] = 'Subject'
msg['From'] = 'vaishalisharmay93@gmail.com'
msg['To'] = 'vikramadityasinghn@gmail.com'
msg.set_content('Hello, this is the content!')


'''
2.)Create an Email Message:

For sending emails, you'll work with an "EmailMessage" object from the "email.message" module.
This object represents the email you want to send.
You can set various attributes like the sender, recipient, subject, and content of the email.
'''
    
server = smtplib.SMTP('smtp.gmail.com', 587)
  
'''
3.)Connect to the SMTP Server:
    

You need to connect to the SMTP server of your email provider. 
For Gmail, the server is 'smtp.gmail.com'. You typically use port '587' for establishing a connection.
'''
server.starttls()
'''
4.) Start a Secure Connection:

Use the starttls() method to initiate a secure connection using TLS (Transport Layer Security).
This is essential for encrypting the communication between your script and the SMTP server.
'''
your_email = 'vaishalisharmay93@gmail.com'
your_password = 'dwaswubatxepcpkb'

server.login(your_email, your_password)
'''
5.) Log in to Your Email Account:

Log in to your email account using the login() method. Provide your email address and password as arguments.
'''
server.quit()


# In[ ]:





# In[ ]:





# In[ ]:




